"""Step 1: extract raw rows from every relevant tab into scratch/extracted_transactions.json.

Deliberately does NO normalization here (no category mapping, no client/retreat
disambiguation) — that's steps 2 and 3. This step's only job is "read the sheet
faithfully and let a human eyeball the result before anything gets clever."

Usage:
    python 01_extract.py /path/to/downloaded-workbook.xlsx

Download the workbook first via File -> Download -> Microsoft Excel (.xlsx)
from the Google Sheet (or File -> Download as .xlsx via the API) — this
script reads a local .xlsx, it doesn't call the Google Sheets API itself.
"""

from __future__ import annotations

import json
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl

from common import (
    MONTH_TAB_NAMES,
    OVERHEAD_LEDGER_TAB,
    SKIP_TABS,
    looks_like_multiple_clients,
)

SCRATCH_DIR = Path(__file__).parent / "scratch"


def is_audit_tab(ws) -> bool:
    """Derived per-retreat rollup tabs (e.g. 'Panther Feb 2026') have a
    'Cash Summary' marker somewhere in columns H-L of the first ~20 rows —
    that's the RETREAT TEMPLATE structure. These are computed FROM the raw
    tabs, not a source themselves; migrating from them too would double-count."""
    for r in range(1, 21):
        for c in range(8, 13):
            v = ws.cell(row=r, column=c).value
            if v and "cash summary" in str(v).lower():
                return True
    return False


def has_header_row(ws) -> dict[str, int] | None:
    """Monthly bank-feed tabs have a real header row with Description/Amount
    columns. Returns a {lowercased header name: 1-based column index} map if
    found, else None."""
    row1 = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    headers = {str(v).strip().lower(): i + 1 for i, v in enumerate(row1) if v}
    if "description" in headers and "amount" in headers:
        # Some tabs (confirmed: MAY, JUNE, JULY) have a blank/whitespace-only
        # header in column 1 instead of literal "ACCOUNT" text, which strips
        # to "" and never matches headers.get("account") below -- silently
        # dropping every account value in the whole tab. Column 1 is always
        # the account column in this tab format regardless of its header
        # text, so fall back to it positionally when nothing else claimed
        # that slot.
        if "account" not in headers and 1 not in headers.values():
            headers["account"] = 1
        return headers
    return None


def to_iso_date(value) -> str | None:
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    if value is None:
        return None
    try:
        return datetime.fromisoformat(str(value)).strftime("%Y-%m-%d")
    except ValueError:
        return None


def extract_header_tab(ws, tab_name: str, headers: dict[str, int]) -> list[dict]:
    """Format A: monthly bank-feed tabs with a real header row."""
    col = {
        "account": headers.get("account"),
        "date": headers.get("date") or headers.get("posting date"),
        "description": headers["description"],
        "amount": headers["amount"],
        "expense_type": headers.get("expense type"),
        "main_category": headers.get("main category"),
        "subcategory": headers.get("expense subcategory"),
        "comment": headers.get("comment"),
    }
    rows = []
    for r in range(2, ws.max_row + 1):
        description = ws.cell(row=r, column=col["description"]).value
        amount = ws.cell(row=r, column=col["amount"]).value
        if description is None and amount is None:
            continue
        rows.append(
            {
                "tab": tab_name,
                "row": r,
                "format": "A_monthly",
                "account": ws.cell(row=r, column=col["account"]).value if col["account"] else None,
                "posted_date": to_iso_date(ws.cell(row=r, column=col["date"]).value) if col["date"] else None,
                "description": str(description).strip() if description is not None else None,
                "amount": amount,
                "expense_type_raw": ws.cell(row=r, column=col["expense_type"]).value if col["expense_type"] else None,
                "main_category_raw": ws.cell(row=r, column=col["main_category"]).value if col["main_category"] else None,
                "subcategory_raw": ws.cell(row=r, column=col["subcategory"]).value if col["subcategory"] else None,
                "retreat_month_raw": to_iso_date(ws.cell(row=r, column=col["comment"]).value) if col["comment"] else None,
            }
        )
    return rows


# Audit tabs built from RETREAT TEMPLATE have a literal template/header row
# disguised as data: a sentinel date (2000-01-01) in column A plus label text
# like "CC Description " / "Expense Type " / "Main Category " in the data
# columns. isinstance-checking column A as a date doesn't filter these out
# (2000-01-01 IS a real date value), so they're excluded explicitly by text.
TEMPLATE_HEADER_STRINGS = {
    "cc description", "description", "expense type", "main category",
    "expense subtype", "expense subcategory", "comment",
}


def extract_sectioned_tab(ws, tab_name: str) -> list[dict]:
    """Format B: legacy per-client tabs — 'Revenue'/'Expenses' section labels
    in column A, positional columns A-G within each section, subtotal rows
    (numeric-only) skipped."""
    rows = []
    section = None
    for r in range(1, ws.max_row + 1):
        col_a = ws.cell(row=r, column=1).value
        if isinstance(col_a, str) and col_a.strip().lower() in ("revenue", "expenses"):
            section = col_a.strip().lower()
            continue

        description = ws.cell(row=r, column=2).value
        amount = ws.cell(row=r, column=3).value
        date_val = col_a
        if description is None or not isinstance(date_val, (datetime, date)):
            continue  # subtotal row, blank row, or something we don't recognize
        if isinstance(description, str) and description.strip().lower() in TEMPLATE_HEADER_STRINGS:
            continue  # RETREAT TEMPLATE's disguised header row (see comment above)

        rows.append(
            {
                "tab": tab_name,
                "row": r,
                "format": "B_legacy_sectioned",
                "section": section,
                "account": None,
                "posted_date": to_iso_date(date_val),
                "description": str(description).strip(),
                "amount": amount,
                "expense_type_raw": ws.cell(row=r, column=4).value,
                "main_category_raw": ws.cell(row=r, column=5).value,
                "subcategory_raw": ws.cell(row=r, column=6).value,
                "retreat_month_raw": to_iso_date(ws.cell(row=r, column=7).value),
            }
        )
    return rows


def extract_overhead_ledger(ws, tab_name: str) -> list[dict]:
    """Format C: the single 2026 Overhead tab — flat rows, no header, no sections."""
    rows = []
    for r in range(1, ws.max_row + 1):
        date_val = ws.cell(row=r, column=1).value
        description = ws.cell(row=r, column=2).value
        amount = ws.cell(row=r, column=3).value
        if not isinstance(date_val, (datetime, date)) or description is None:
            continue
        rows.append(
            {
                "tab": tab_name,
                "row": r,
                "format": "C_overhead_ledger",
                "account": None,
                "posted_date": to_iso_date(date_val),
                "description": str(description).strip(),
                "amount": amount,
                "expense_type_raw": ws.cell(row=r, column=4).value,
                "main_category_raw": ws.cell(row=r, column=5).value,
                "subcategory_raw": ws.cell(row=r, column=6).value,
                "retreat_month_raw": None,  # overhead has no retreat-month concept
                "note": ws.cell(row=r, column=8).value,
            }
        )
    return rows


def row_key(row: dict) -> tuple:
    """Dedup key for detecting whether an audit-tab row is already captured
    by a primary source (monthly/legacy tab) — same date + same amount is a
    strong enough signal at this data volume (collisions would need two
    transactions of the identical dollar amount on the identical day, which
    a human reviewing scratch/ output would still catch)."""
    amount = row.get("amount")
    try:
        amount = round(float(amount), 2)
    except (TypeError, ValueError):
        amount = None
    return (row.get("posted_date"), amount)


def main(workbook_path: str):
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    SCRATCH_DIR.mkdir(exist_ok=True)

    all_rows: list[dict] = []
    ambiguous_multi_client: list[dict] = []
    audit_tab_names: list[str] = []
    unrecognized_tabs: list[str] = []

    for tab_name in wb.sheetnames:
        if tab_name in SKIP_TABS:
            continue
        ws = wb[tab_name]

        if tab_name == OVERHEAD_LEDGER_TAB:
            all_rows.extend(extract_overhead_ledger(ws, tab_name))
            continue

        if tab_name in MONTH_TAB_NAMES:
            headers = has_header_row(ws)
            if headers:
                all_rows.extend(extract_header_tab(ws, tab_name, headers))
                continue

        if is_audit_tab(ws):
            # Don't extract yet — audit tabs are USUALLY fully redundant with
            # a monthly/legacy tab (filtered copies), but some retreats (seen
            # concretely: several October 2026 retreats) have NO monthly-tab
            # data at all because that month's raw bank tab was never
            # populated — for those, the audit tab is the only copy of real
            # revenue/COGS data that exists. Resolved below via a row-level
            # dedup pass against everything already extracted, once all
            # primary sources have been read.
            audit_tab_names.append(tab_name)
            continue

        headers = has_header_row(ws)
        if headers:
            all_rows.extend(extract_header_tab(ws, tab_name, headers))
            continue

        # Try the sectioned format; if it yields nothing, flag for a human look.
        sectioned_rows = extract_sectioned_tab(ws, tab_name)
        if sectioned_rows:
            all_rows.extend(sectioned_rows)
        else:
            unrecognized_tabs.append(tab_name)

    # Fallback pass: for each audit/rollup tab, extract its rows the same way
    # as a legacy sectioned tab (they share the same A-G column layout — the
    # audit-specific Cash Summary/rollup content lives in columns H+ and is
    # ignored here), then keep only rows whose (date, amount) aren't already
    # present from a primary source. A tab that's fully redundant contributes
    # nothing; a tab that's the only source for its retreat contributes
    # everything, tagged fallback_source so it's visible in the review step.
    primary_keys = {row_key(r) for r in all_rows}
    fallback_rows: list[dict] = []
    redundant_audit_tabs: list[str] = []
    fallback_source_tabs: list[str] = []

    for tab_name in audit_tab_names:
        ws = wb[tab_name]
        candidate_rows = extract_sectioned_tab(ws, tab_name)
        novel = [r for r in candidate_rows if row_key(r) not in primary_keys]
        if not candidate_rows:
            continue
        if len(novel) / len(candidate_rows) < 0.5:
            # Mostly overlaps a primary source — treat as the redundant
            # derived copy it almost certainly is, and skip entirely (partial
            # overlap here is more likely two similar-but-different charges
            # than a real gap, and taking the whole tab would double-count
            # the overlapping majority).
            redundant_audit_tabs.append(tab_name)
            continue
        for row in novel:
            row["fallback_source"] = True
        fallback_rows.extend(novel)
        fallback_source_tabs.append(tab_name)

    all_rows.extend(fallback_rows)

    # Flag (but still include) rows whose subcategory names multiple clients —
    # v1 has no transaction-splitting, so these need a human decision.
    for row in all_rows:
        sub = row.get("subcategory_raw")
        if isinstance(sub, str) and looks_like_multiple_clients(sub):
            ambiguous_multi_client.append(row)

    (SCRATCH_DIR / "extracted_transactions.json").write_text(
        json.dumps(all_rows, indent=2, default=str)
    )
    (SCRATCH_DIR / "ambiguous_multi_client_rows.json").write_text(
        json.dumps(ambiguous_multi_client, indent=2, default=str)
    )
    (SCRATCH_DIR / "redundant_audit_tabs_skipped.txt").write_text("\n".join(sorted(redundant_audit_tabs)))
    (SCRATCH_DIR / "fallback_source_audit_tabs.txt").write_text("\n".join(sorted(fallback_source_tabs)))
    (SCRATCH_DIR / "unrecognized_tabs.txt").write_text("\n".join(sorted(unrecognized_tabs)))

    print(f"Extracted {len(all_rows)} rows from {len(wb.sheetnames)} tabs.")
    print(f"  {len(redundant_audit_tabs)} derived audit/rollup tabs skipped as fully redundant "
          f"(see scratch/redundant_audit_tabs_skipped.txt).")
    if fallback_source_tabs:
        print(f"  {len(fallback_rows)} rows recovered from {len(fallback_source_tabs)} audit tabs that "
              f"had NO corresponding monthly/legacy source data (see scratch/fallback_source_audit_tabs.txt) "
              f"— these are tagged \"fallback_source\": true in the JSON. Spot-check these specifically; "
              f"the 50% novelty threshold used to decide 'fallback' vs 'redundant' is a heuristic, not a guarantee.")
    if unrecognized_tabs:
        print(f"  WARNING: {len(unrecognized_tabs)} tabs didn't match any known format — "
              f"see scratch/unrecognized_tabs.txt. Nothing was extracted from these.")
    if ambiguous_multi_client:
        print(f"  WARNING: {len(ambiguous_multi_client)} rows look like shared/split charges "
              f"across multiple clients — see scratch/ambiguous_multi_client_rows.json. "
              f"These need manual assignment before loading (v1 has no transaction splitting).")
    print(f"\nReview scratch/extracted_transactions.json before running 02_category_map.py.")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python 01_extract.py /path/to/workbook.xlsx")
        sys.exit(1)
    main(sys.argv[1])
