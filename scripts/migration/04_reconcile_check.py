"""Step 4 (after 03_load.py has run): sanity-check migrated totals against
the sheet's own '2025 Summary' / '2026 Summary' tabs. Those tabs' "Final
Revenue" figures are themselves formulas summing the sheet's per-retreat
audit tabs, so they're a ready-made ground truth to diff against — this is
the single most useful automated correctness check available.

This does NOT block anything or modify data; it only prints a variance
report. The sheet's own totals could themselves be stale (some retreats had
blank "Final Revenue" cells at the time this was written, e.g. most of
Oct/Nov 2026 — see the migration README), so treat large/systematic
mismatches as worth investigating, not proof of a bug.

Usage:
    python 04_reconcile_check.py /path/to/workbook.xlsx
"""

from __future__ import annotations

import os
import sys
from pathlib import Path

import openpyxl
from dotenv import load_dotenv
from supabase import create_client

from common import canonical_client_name, split_client_and_retreat_name


def get_supabase():
    load_dotenv(Path(__file__).parent.parent.parent / ".env.local")
    url = os.environ.get("NEXT_PUBLIC_SUPABASE_URL")
    key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    if not url or not key:
        raise SystemExit("NEXT_PUBLIC_SUPABASE_URL / SUPABASE_SERVICE_ROLE_KEY not found in .env.local")
    return create_client(url, key)


MONTH_NAME_TO_NUM = {
    "january": 1, "february": 2, "march": 3, "april": 4, "may": 5, "june": 6,
    "july": 7, "august": 8, "september": 9, "october": 10, "november": 11, "december": 12,
}


def sheet_summary_rows(wb, tab_name: str, year: int, revenue_col: int) -> list[dict]:
    """revenue_col differs by tab: "2025 Summary" has columns
    Name/Month/Revenue/Gross Profit/Margin/.../Cash Position (no quoted-vs-
    actual split, so column 2 IS the actual revenue); "2026 Summary" has
    Name/Month/Revenue at Pricing/Final Revenue/... (column 3 is the actual
    "Final Revenue"). Using the same index for both silently compared our
    migrated revenue against Gross Profit for every 2025 row."""
    if tab_name not in wb.sheetnames:
        return []
    ws = wb[tab_name]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[0]
        month = row[1]
        final_revenue = row[revenue_col] if len(row) > revenue_col else None
        if not name or not month:
            continue
        month_num = MONTH_NAME_TO_NUM.get(str(month).strip().lower())
        if month_num is None:
            continue
        retreat_month = f"{year:04d}-{month_num:02d}-01"
        rows.append({
            "name": str(name).strip(),
            "month": str(month).strip(),
            "retreat_month": retreat_month,
            "final_revenue": final_revenue,
        })
    return rows


def main(workbook_path: str):
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    supabase = get_supabase()

    sheet_rows = (
        sheet_summary_rows(wb, "2025 Summary", 2025, revenue_col=2)
        + sheet_summary_rows(wb, "2026 Summary", 2026, revenue_col=3)
    )

    retreats = (
        supabase.table("retreats")
        .select("id, name, retreat_month, client:clients(name)")
        .execute()
        .data
    )
    actuals = supabase.table("retreat_actuals").select("retreat_id, revenue").execute().data
    revenue_by_retreat_id = {a["retreat_id"]: a["revenue"] for a in actuals}

    matched = 0
    mismatched = []
    no_migrated_data = []

    for sheet_row in sheet_rows:
        if sheet_row["final_revenue"] in (None, "", " "):
            continue  # sheet itself has no final figure yet — nothing to compare

        client_name, retreat_name = split_client_and_retreat_name(sheet_row["name"])
        client_name = canonical_client_name(client_name)
        retreat_name = canonical_client_name(retreat_name)  # same fix as 02_category_map.py

        candidate = next(
            (
                r for r in retreats
                if r["name"].strip() == retreat_name
                and (r["client"]["name"] if r["client"] else "") == client_name
                and r["retreat_month"] == sheet_row["retreat_month"]
            ),
            None,
        )
        if not candidate:
            no_migrated_data.append(sheet_row)
            continue

        migrated_revenue = revenue_by_retreat_id.get(candidate["id"], 0)
        sheet_revenue = float(sheet_row["final_revenue"])
        if abs(migrated_revenue - sheet_revenue) < 1.0:
            matched += 1
        else:
            mismatched.append(
                {**sheet_row, "migrated_revenue": migrated_revenue, "variance": migrated_revenue - sheet_revenue}
            )

    print(f"{matched} retreats match the sheet's own Final Revenue within $1.")
    print(f"{len(mismatched)} retreats have a variance — review each:")
    for m in mismatched:
        print(f"  {m['name']} / {m['month']}: sheet={m['final_revenue']}, "
              f"migrated={m['migrated_revenue']}, variance={m['variance']:+.2f}")
    print(f"{len(no_migrated_data)} sheet entries had no matching migrated retreat found "
          f"(name/month didn't resolve to the same client+retreat — check client_retreat_map.csv "
          f"from step 2, or the retreat genuinely wasn't migrated):")
    for n in no_migrated_data:
        print(f"  {n['name']} / {n['month']}")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python 04_reconcile_check.py /path/to/workbook.xlsx")
        sys.exit(1)
    main(sys.argv[1])
